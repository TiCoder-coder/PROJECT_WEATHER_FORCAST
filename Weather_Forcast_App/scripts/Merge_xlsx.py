import sys
import re
from pathlib import Path
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet


# ============================================================
# FIX ENCODING CHO WINDOWS CONSOLE (chống lỗi in tiếng Việt)
# ============================================================
# - Trên Windows, terminal đôi khi dùng encoding mặc định không phải UTF-8
# - Khi print tiếng Việt dễ bị lỗi/ra ký tự lạ
# - Đoạn này bọc lại stdout/stderr để luôn dùng UTF-8
if sys.platform == "win32":
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8")


# ============================================================
# CẤU HÌNH THƯ MỤC & TÊN FILE
# ============================================================
# DYNAMIC PATH: tự tính từ vị trí project root, không hardcode Linux path
_PROJECT_ROOT = str(Path(__file__).resolve().parents[2])
MERGE_DIR_ABS = str(Path(_PROJECT_ROOT) / "data" / "data_merge")
OUTPUT_DIR_ABS = str(Path(_PROJECT_ROOT) / "data" / "data_crawl")

# Tên file output sau khi merge cho nhóm "khác" (không phải vietnam_weather_)
MERGE_FILENAME = "merged_vrain_data.xlsx"

# Tên file output sau khi merge cho nhóm vietnam_weather_
MERGE_VIETNAM_FILENAME = "merged_vietnam_weather_data.xlsx"

# Log file dùng để lưu danh sách tên file đã merge (tránh merge lại lần sau)
LOG_FILENAME = "merged_files_log.txt"
LOG_VIETNAM_FILENAME = "merged_vietnam_files_log.txt"

# ============================================================
# CƠ CHẾ SAVE THEO BATCH (tránh mất dữ liệu, giảm rủi ro crash)
# ============================================================
# - Khi append quá nhiều dòng vào workbook, nếu crash giữa chừng sẽ mất công
# - SAVE_EVERY_N_ROWS giúp cứ mỗi N dòng append sẽ save 1 lần xuống disk
# - N càng nhỏ càng an toàn nhưng save nhiều -> chậm hơn
SAVE_EVERY_N_ROWS = 30000

MASTER_COLUMNS = [
    "station_id",
    "station_name",
    "province",
    "district",
    "latitude",
    "longitude",
    "timestamp",
    "data_source",
    "data_quality",
    "data_time",
    "temperature_current",
    "temperature_max",
    "temperature_min",
    "temperature_avg",
    "humidity_current",
    "humidity_max",
    "humidity_min",
    "humidity_avg",
    "pressure_current",
    "pressure_max",
    "pressure_min",
    "pressure_avg",
    "wind_speed_current",
    "wind_speed_max",
    "wind_speed_min",
    "wind_speed_avg",
    "wind_direction_current",
    "wind_direction_avg",
    "rain_current",
    "rain_max",
    "rain_min",
    "rain_avg",
    "rain_total",
    "cloud_cover_current",
    "cloud_cover_max",
    "cloud_cover_min",
    "cloud_cover_avg",
    "visibility_current",
    "visibility_max",
    "visibility_min",
    "visibility_avg",
    "thunder_probability",
    "status",
]

# Mapping chuẩn hóa từ tiếng Việt, camelCase, snake_case cũ sang snake_case chuẩn
COLUMN_SCHEMA_MAPPING = {
    # Tiếng Việt sang snake_case
    "Mã trạm": "station_id",
    "Tên trạm": "station_name",
    "Tỉnh/Thành phố": "province",
    "Huyện": "district",
    "Vĩ độ": "latitude",
    "Kinh độ": "longitude",
    "Dấu thời gian": "timestamp",
    "Nguồn dữ liệu": "data_source",
    "Chất lượng dữ liệu": "data_quality",
    "Thời gian cập nhật": "data_time",
    "Nhiệt độ hiện tại": "temperature_current",
    "Nhiệt độ tối đa": "temperature_max",
    "Nhiệt độ tối thiểu": "temperature_min",
    "Nhiệt độ trung bình": "temperature_avg",
    "Độ ẩm hiện tại": "humidity_current",
    "Độ ẩm tối đa": "humidity_max",
    "Độ ẩm tối thiểu": "humidity_min",
    "Độ ẩm trung bình": "humidity_avg",
    "Áp suất hiện tại": "pressure_current",
    "Áp suất tối đa": "pressure_max",
    "Áp suất tối thiểu": "pressure_min",
    "Áp suất trung bình": "pressure_avg",
    "Tốc độ gió hiện tại": "wind_speed_current",
    "Tốc độ gió tối đa": "wind_speed_max",
    "Tốc độ gió tối thiểu": "wind_speed_min",
    "Tốc độ gió trung bình": "wind_speed_avg",
    "Hướng gió hiện tại": "wind_direction_current",
    "Hướng gió trung bình": "wind_direction_avg",
    "Lượng mưa hiện tại": "rain_current",
    "Lượng mưa tối đa": "rain_max",
    "Lượng mưa tối thiểu": "rain_min",
    "Lượng mưa trung bình": "rain_avg",
    "Tổng lượng mưa": "rain_total",
    "Độ che phủ mây hiện tại": "cloud_cover_current",
    "Độ che phủ mây tối đa": "cloud_cover_max",
    "Độ che phủ mây tổi thiểu": "cloud_cover_min",
    "Độ che phủ mây trung bình": "cloud_cover_avg",
    "Tầm nhìn hiện tại": "visibility_current",
    "Tầm nhìn đa": "visibility_max",
    "Tầm nhìn tối thiểu": "visibility_min",
    "Tầm nhìn trung bình": "visibility_avg",
    "Xác xuất sấm sét": "thunder_probability",
    "Tình trạng": "status",
    # camelCase sang snake_case
    "stationId": "station_id",
    "stationName": "station_name",
    "dataSource": "data_source",
    "dataQuality": "data_quality",
    "dataTime": "data_time",
    "temperatureCurrent": "temperature_current",
    "temperatureMax": "temperature_max",
    "temperatureMin": "temperature_min",
    "temperatureAvg": "temperature_avg",
    "humidityCurrent": "humidity_current",
    "humidityMax": "humidity_max",
    "humidityMin": "humidity_min",
    "humidityAvg": "humidity_avg",
    "pressureCurrent": "pressure_current",
    "pressureMax": "pressure_max",
    "pressureMin": "pressure_min",
    "pressureAvg": "pressure_avg",
    "windSpeedCurrent": "wind_speed_current",
    "windSpeedMax": "wind_speed_max",
    "windSpeedMin": "wind_speed_min",
    "windSpeedAvg": "wind_speed_avg",
    "windDirectionCurrent": "wind_direction_current",
    "windDirectionAvg": "wind_direction_avg",
    "rainCurrent": "rain_current",
    "rainMax": "rain_max",
    "rainMin": "rain_min",
    "rainAvg": "rain_avg",
    "rainTotal": "rain_total",
    "cloudCoverCurrent": "cloud_cover_current",
    "cloudCoverMax": "cloud_cover_max",
    "cloudCoverMin": "cloud_cover_min",
    "cloudCoverAvg": "cloud_cover_avg",
    "visibilityCurrent": "visibility_current",
    "visibilityMax": "visibility_max",
    "visibilityMin": "visibility_min",
    "visibilityAvg": "visibility_avg",
    "thunderProbability": "thunder_probability",
    # snake_case cũ sang snake_case chuẩn (giữ nguyên)
    "station_id": "station_id",
    "station_name": "station_name",
    "data_source": "data_source",
    "data_quality": "data_quality",
    "data_time": "data_time",
    "temperature_current": "temperature_current",
    "temperature_max": "temperature_max",
    "temperature_min": "temperature_min",
    "temperature_avg": "temperature_avg",
    "humidity_current": "humidity_current",
    "humidity_max": "humidity_max",
    "humidity_min": "humidity_min",
    "humidity_avg": "humidity_avg",
    "pressure_current": "pressure_current",
    "pressure_max": "pressure_max",
    "pressure_min": "pressure_min",
    "pressure_avg": "pressure_avg",
    "wind_speed_current": "wind_speed_current",
    "wind_speed_max": "wind_speed_max",
    "wind_speed_min": "wind_speed_min",
    "wind_speed_avg": "wind_speed_avg",
    "wind_direction_current": "wind_direction_current",
    "wind_direction_avg": "wind_direction_avg",
    "rain_current": "rain_current",
    "rain_max": "rain_max",
    "rain_min": "rain_min",
    "rain_avg": "rain_avg",
    "rain_total": "rain_total",
    "cloud_cover_current": "cloud_cover_current",
    "cloud_cover_max": "cloud_cover_max",
    "cloud_cover_min": "cloud_cover_min",
    "cloud_cover_avg": "cloud_cover_avg",
    "visibility_current": "visibility_current",
    "visibility_max": "visibility_max",
    "visibility_min": "visibility_min",
    "visibility_avg": "visibility_avg",
    "thunder_probability": "thunder_probability",
}

# ============================================================
# COLUMN_ALIASES = BẢN ĐỒ CHUẨN HOÁ TÊN CỘT (alias -> chuẩn)
# ============================================================
# - Dữ liệu nguồn có thể bị gõ sai dấu hoặc khác cách viết
# - Ví dụ "tối thiểu" bị viết nhầm thành "tổi thiểu"
# - Ở đây bạn quyết định dùng "Độ che phủ mây tổi thiểu" làm dạng chuẩn
#   => mọi biến thể khác sẽ được map về dạng này
# - Điều này giúp tránh sinh ra 2 cột gần giống nhau chỉ vì khác chữ
COLUMN_ALIASES = {
    "Độ che phủ mây tối thiểu": "Độ che phủ mây tổi thiểu",
    "Tầm nhìn tối đa": "Tầm nhìn đa",
    "Xác suất sấm sét": "Xác xuất sấm sét",
    "Xác suất sét": "Xác xuất sấm sét",
    "Xác suất sấm sét": "Xác xuất sấm sét",
}


def norm_col(x) -> str:
    # ============================================================
    # CHUẨN HOÁ TÊN CỘT
    # ============================================================
    # - Ép về string + strip() để bỏ khoảng trắng đầu/cuối
    # - regex thay nhiều khoảng trắng liên tiếp thành 1 khoảng trắng
    # - Sau đó tra COLUMN_ALIASES:
    #   + nếu có alias -> trả về tên chuẩn
    #   + nếu không -> giữ nguyên
    s = re.sub(r"\s+", " ", str(x).strip())
    return COLUMN_ALIASES.get(s, s)


def load_processed_files(log_path: Path) -> set[str]:
    # ============================================================
    # ĐỌC LOG -> TẬP TÊN FILE ĐÃ XỬ LÝ
    # ============================================================
    # - Mục tiêu: chạy script nhiều lần nhưng không merge lại file cũ
    # - log file chỉ chứa mỗi dòng là tên file .xlsx đã merge
    # - Nếu log chưa tồn tại -> trả set rỗng
    if not log_path.exists():
        return set()
    processed = set()
    try:
        with log_path.open("r", encoding="utf-8") as f:
            for line in f:
                name = line.strip()
                if name:
                    processed.add(name)
    except Exception as e:
        # Nếu log lỗi cũng không cho crash toàn bộ merge
        print(f"Loi khi doc log file {log_path.name}: {e}")
    return processed


def save_processed_files(log_path: Path, processed_files: set[str]) -> None:
    # ============================================================
    # GHI LOG: LƯU DANH SÁCH FILE ĐÃ MERGE
    # ============================================================
    # - Ghi toàn bộ set processed_files ra file
    # - sorted(...) để log ổn định, dễ đọc/diff
    try:
        with log_path.open("w", encoding="utf-8") as f:
            for name in sorted(processed_files):
                f.write(name + "\n")
    except Exception as e:
        # Không crash nếu log ghi lỗi
        print(f"Loi khi ghi log file {log_path.name}: {e}")


def read_excel_file(file_path: Path) -> pd.DataFrame:
    # ============================================================
    # ĐỌC 1 FILE EXCEL -> DATAFRAME
    # ============================================================
    # - pd.read_excel sẽ đọc sheet đầu tiên mặc định
    # - Nếu file lỗi/đọc không được -> trả DataFrame rỗng
    try:
        df = pd.read_excel(file_path)
        if df is None or df.empty:
            return pd.DataFrame()
        return df
    except Exception as e:
        print(f"Loi khi doc file {file_path.name}: {e}")
        return pd.DataFrame()


def clean_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    # ============================================================
    # LÀM SẠCH DATAFRAME TRƯỚC KHI MERGE
    # ============================================================
    # Những việc làm chính:
    # 1) copy() để tránh sửa trực tiếp df gốc
    # 2) Chuẩn hoá tên cột bằng norm_col
    # 3) Loại bỏ cột trùng tên sau chuẩn hoá (duplicated)
    # 4) Xoá các cột "Unnamed: ..." (thường do excel xuất dư cột)
    df = df.copy()

    # Chuẩn hoá tên cột và mapping về snake_case chuẩn
    df.columns = [COLUMN_SCHEMA_MAPPING.get(norm_col(c), norm_col(c)) for c in df.columns]

    # Nếu sau chuẩn hoá bị trùng cột (vd: 2 cột khác nhau map về cùng alias)
    # -> giữ cột đầu tiên, bỏ cột trùng
    df = df.loc[:, ~df.columns.duplicated()]

    # Các cột kiểu "Unnamed: 0", "Unnamed: 1" thường do index/format excel
    drop_cols = [c for c in df.columns if str(c).lower().startswith("unnamed")]
    if drop_cols:
        df = df.drop(columns=drop_cols, errors="ignore")

    return df


def get_new_excel_files(output_dir: Path, processed_files: set[str]) -> tuple[list[Path], list[Path]]:
    # ============================================================
    # QUÉT THƯ MỤC output -> LẤY CÁC FILE .xlsx CHƯA XỬ LÝ
    # ============================================================
    # - output_dir: nơi chứa file excel nguồn
    # - processed_files: tập tên file đã merge (từ log)
    #
    # Trả về 2 list:
    # 1) vietnam_files: file có prefix "vietnam_weather_"
    # 2) other_files: các file còn lại
    if not output_dir.exists():
        print(f"Thu muc nguon khong ton tai: {output_dir}")
        return [], []

    all_excel_files = sorted(output_dir.glob("*.xlsx"))
    if not all_excel_files:
        print(f"Khong tim thay file .xlsx nao trong thu muc: {output_dir}")
        return [], []

    vietnam_files, other_files = [], []
    for file_path in all_excel_files:
        # Nếu file đã nằm trong log -> bỏ qua
        if file_path.name in processed_files:
            continue

        # Phân loại theo prefix tên file
        if file_path.name.startswith("vietnam_weather_"):
            vietnam_files.append(file_path)
        else:
            other_files.append(file_path)

    # In thống kê để biết pipeline đang xử lý được bao nhiêu file mới
    print(f"Tong so file .xlsx trong output: {len(all_excel_files)}")
    print(f"So file vietnam_weather_ moi: {len(vietnam_files)}")
    print(f"So file khac moi: {len(other_files)}")
    return vietnam_files, other_files


def _load_or_create_wb_ws(merge_path: Path) -> tuple[Workbook, Worksheet, list[str]]:
    # ============================================================
    # MỞ FILE MERGE NẾU ĐÃ TỒN TẠI, HOẶC TẠO FILE MỚI
    # ============================================================
    # Trả về:
    # - wb: Workbook openpyxl
    # - ws: worksheet active
    # - header: list tên cột hiện có trong dòng 1 của file merge
    #
    # Ý tưởng:
    # - Nếu merge_path tồn tại:
    #   + load_workbook để tiếp tục append vào file cũ
    #   + đọc header row (row=1) để biết schema hiện tại
    # - Nếu không tồn tại:
    #   + tạo workbook mới và viết MASTER_COLUMNS làm header
    if merge_path.exists():
        wb = load_workbook(merge_path)
        ws = wb.active

        # Đọc header dòng 1 theo số cột hiện tại
        header = []
        max_col = ws.max_column
        for col in range(1, max_col + 1):
            v = ws.cell(row=1, column=col).value
            if v is None:
                continue
            header.append(norm_col(v))

        # Nếu file merge có nhưng dòng header bị trống/hỏng
        # -> reset lại header theo MASTER_COLUMNS
        if not header:
            header = list(MASTER_COLUMNS)
            for i, col_name in enumerate(header, start=1):
                ws.cell(row=1, column=i).value = col_name

        return wb, ws, header

    # Nếu file merge chưa tồn tại -> tạo mới
    wb = Workbook()
    ws = wb.active
    ws.title = "data"

    # Ghi header = MASTER_COLUMNS vào dòng 1
    header = list(MASTER_COLUMNS)
    for i, col_name in enumerate(header, start=1):
        ws.cell(row=1, column=i).value = col_name

    # Save ngay để tạo file vật lý trên disk
    wb.save(merge_path)
    return wb, ws, header


def _ensure_header_has_columns(ws: Worksheet, header: list[str], cols_to_ensure: list[str]) -> list[str]:
    # ============================================================
    # ĐẢM BẢO HEADER CÓ ĐỦ CÁC CỘT TRONG cols_to_ensure
    # ============================================================
    # - Nếu thiếu cột -> append vào cuối header + ghi vào row 1 của sheet
    # - Trả về header mới (có thể đã được mở rộng)
    changed = False
    for c in cols_to_ensure:
        c = norm_col(c)
        if c not in header:
            header.append(c)
            ws.cell(row=1, column=len(header)).value = c
            changed = True
    if changed:
        print(f"  + Da mo rong schema, tong so cot hien tai: {len(header)}")
    return header


def _to_excel_value(v):
    # ============================================================
    # CHUYỂN GIÁ TRỊ PANDAS -> GIÁ TRỊ HỢP LỆ CHO OPENPYXL
    # ============================================================
    # - pd.NA / NaN: openpyxl không thích -> đổi thành None
    # - pd.Timestamp: đổi sang datetime python bằng to_pydatetime()
    # - còn lại giữ nguyên
    try:
        if pd.isna(v):
            return None
    except Exception:
        # Một số kiểu object có thể khiến pd.isna lỗi -> bỏ qua
        pass
    if isinstance(v, pd.Timestamp):
        return v.to_pydatetime()
    return v


def append_df_incremental(
    wb: Workbook,
    ws: Worksheet,
    header: list[str],
    df: pd.DataFrame,
    merge_path: Path,
    save_every_n_rows: int = SAVE_EVERY_N_ROWS,
) -> int:
    # ============================================================
    # APPEND DATAFRAME VÀO FILE EXCEL THEO KIỂU INCREMENTAL
    # ============================================================
    # Mục tiêu:
    # - Append từng dòng một vào worksheet để không phải giữ toàn bộ dữ liệu trong RAM
    # - Save định kỳ mỗi save_every_n_rows dòng để:
    #   + giảm rủi ro mất dữ liệu
    #   + tránh workbook quá lớn mà chưa save
    #
    # Bước chính:
    # 1) Bổ sung các cột còn thiếu trong df theo header (NA)
    # 2) Reorder df theo đúng thứ tự header
    # 3) itertuples để iterate nhanh hơn iterrows
    # 4) ws.append từng dòng
    # 5) save theo batch
    for c in header:
        # Nếu df không có cột nào đó trong schema -> tạo cột đó toàn NA
        if c not in df.columns:
            df[c] = pd.NA
    # Chỉ giữ đúng các cột trong header (không mở rộng thêm cột ngoài schema)
    df = df[header]

    appended = 0
    buffer_count = 0

    # itertuples(index=False, name=None) -> trả tuple thuần, nhanh
    for row in df.itertuples(index=False, name=None):
        # Convert giá trị sang dạng excel-friendly
        excel_row = [_to_excel_value(x) for x in row]
        ws.append(excel_row)
        appended += 1
        buffer_count += 1

        # Save định kỳ để giảm rủi ro crash/mất dữ liệu
        if save_every_n_rows > 0 and buffer_count >= save_every_n_rows:
            wb.save(merge_path)
            buffer_count = 0

    # Save cuối cùng sau khi append xong
    wb.save(merge_path)
    return appended


def merge_single_category_incremental(
    file_list: list[Path],
    merge_path: Path,
    log_path: Path,
    processed_files: set[str],
    category_name: str,
) -> None:
    # ============================================================
    # MERGE 1 NHÓM FILE (category) THEO KIỂU INCREMENTAL
    # ============================================================
    # Tham số:
    # - file_list: danh sách file cần merge (chưa processed)
    # - merge_path: file excel đích để append vào
    # - log_path: file log của category đó
    # - processed_files: set tên file đã xử lý (riêng cho category)
    # - category_name: tên hiển thị/log (vd: "vietnam_weather_" hoặc "khac")
    #
    # Luồng xử lý:
    # - Mở hoặc tạo file merge
    # - Đảm bảo header có MASTER_COLUMNS
    # - Với mỗi file:
    #   + đọc -> clean
    #   + nếu phát hiện cột mới -> mở rộng header
    #   + append dữ liệu incremental
    #   + ghi log ngay sau khi append thành công
    if not file_list:
        print(f"Khong co file {category_name} moi de merge.")
        return

    print(f"\n=== MERGE INCREMENTAL: {category_name.upper()} ===")
    print(f"So luong file: {len(file_list)}")
    print(f"File merge: {merge_path}")

    # Mở workbook merge hiện có hoặc tạo mới nếu chưa có
    wb, ws, header = _load_or_create_wb_ws(merge_path)

    # Đảm bảo header chứa đầy đủ MASTER_COLUMNS (schema chuẩn)
    header = _ensure_header_has_columns(ws, header, MASTER_COLUMNS)
    wb.save(merge_path)

    ok_count = 0
    for idx, file_path in enumerate(sorted(file_list), start=1):
        print(f"\n[{idx}/{len(file_list)}] Dang xu ly: {file_path.name}")

        # Đọc file excel nguồn
        df = read_excel_file(file_path)
        if df.empty:
            print("  - File rong/khong doc duoc, bo qua.")
            continue

        # Làm sạch: chuẩn hoá tên cột, bỏ cột unnamed, bỏ cột trùng
        df = clean_dataframe(df)

        # Không mở rộng schema, chỉ giữ đúng các cột chuẩn
        new_cols = [c for c in df.columns if c not in header]
        if new_cols:
            print(f"  ! File có cột lạ không thuộc schema chuẩn, sẽ bị loại bỏ: {new_cols}")

        try:
            # Append dữ liệu theo từng dòng + save theo batch
            appended = append_df_incremental(
                wb=wb,
                ws=ws,
                header=header,
                df=df,
                merge_path=merge_path,
                save_every_n_rows=SAVE_EVERY_N_ROWS,
            )
            print(f"  ✓ Da append {appended} dong tu {file_path.name}")

            # Sau khi append thành công:
            # - đánh dấu file đã xử lý
            # - ghi log ngay để nếu script dừng đột ngột vẫn không merge lại file này
            processed_files.add(file_path.name)
            save_processed_files(log_path, processed_files)
            ok_count += 1
        except Exception as e:
            # Nếu append lỗi:
            # - in lỗi
            # - KHÔNG ghi log (để lần sau chạy lại file này)
            print(f"  ✗ Loi khi append file {file_path.name}: {e}")
            print("  -> Khong danh dau processed (de lan sau chay lai).")

    # Save cuối cùng cho chắc chắn
    wb.save(merge_path)
    print(f"\n=== XONG {category_name}: OK {ok_count}/{len(file_list)} file ===")


def merge_excel_files_once(base_dir: Path) -> None:
    # ============================================================
    # HÀM "ĐIỀU PHỐI" MERGE CHÍNH (CHẠY 1 LẦN)
    # ============================================================
    # - base_dir: thư mục gốc (thường là root project)
    #
    # Tạo cấu trúc:
    # base_dir/output       : file excel nguồn
    # base_dir/Merge_data   : file merged + logs
    #
    # Merge theo 2 nhóm:
    # 1) vietnam_weather_  -> merged_vietnam_weather_data.xlsx + log riêng
    # 2) các file còn lại  -> merged_vrain_data.xlsx + log riêng

    output_dir = Path(OUTPUT_DIR_ABS)
    merge_dir = Path(MERGE_DIR_ABS)
    merge_dir.mkdir(parents=True, exist_ok=True)

    merge_vietnam_path = merge_dir / MERGE_VIETNAM_FILENAME
    log_vietnam_path = merge_dir / LOG_VIETNAM_FILENAME

    merge_other_path = merge_dir / MERGE_FILENAME
    log_other_path = merge_dir / LOG_FILENAME

    print("======== BAT DAU MERGE =========")
    print(f"Thu muc nguon (output):     {output_dir}")
    print(f"Thu muc merge (Merge_data): {merge_dir}")
    print("================================")

    # Đọc log đã xử lý của từng nhóm
    processed_vietnam = load_processed_files(log_vietnam_path)
    processed_other = load_processed_files(log_other_path)

    # Hợp 2 tập lại để lọc tất cả file đã xử lý (không phân biệt nhóm)
    processed_all = processed_vietnam.union(processed_other)

    # Lấy danh sách file mới trong output (chưa nằm trong processed_all)
    vietnam_files, other_files = get_new_excel_files(output_dir, processed_all)

    # Merge nhóm vietnam_weather_
    merge_single_category_incremental(
        vietnam_files,
        merge_vietnam_path,
        log_vietnam_path,
        processed_vietnam,
        "vietnam_weather_"
    )

    # Merge nhóm còn lại
    merge_single_category_incremental(
        other_files,
        merge_other_path,
        log_other_path,
        processed_other,
        "khac"
    )

    print("\n======== KET THUC MERGE =========")


if __name__ == "__main__":
    # ============================================================
    # ENTRY POINT (khi chạy file python trực tiếp)
    # ============================================================
    # - SCRIPT_DIR: thư mục chứa file script hiện tại
    # - BASE_DIR: thư mục cha của SCRIPT_DIR (thường là root project)
    # Không cần BASE_DIR nữa, dùng đường dẫn tuyệt đối
    output_dir = Path(OUTPUT_DIR_ABS)
    if not output_dir.exists():
        print(f"ERROR: Khong tim thay thu muc output tai: {output_dir}")
        sys.exit(1)

    # Chạy merge 1 lần
    merge_excel_files_once(None)
